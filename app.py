#!/usr/bin/env python3
"""
ABET CPE Course Mapping Desktop Application
Native desktop app for managing course-to-ABET learning outcome mappings
"""

from flask import Flask, render_template, jsonify, request
import json
import os
import subprocess
import sys
from contextlib import contextmanager
try:
    import psycopg
except Exception:
    psycopg = None
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Get the base path for resources (works with PyInstaller)
def get_resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        # Use the app file location so hosted runtimes don't depend on cwd.
        base_path = os.path.dirname(os.path.abspath(__file__))
    
    return os.path.join(base_path, relative_path)

# Get the directory where the executable/script is located (for writable files)
def get_data_dir():
    """Get directory for writable data files"""
    if getattr(sys, 'frozen', False):
        # Running as compiled executable
        return os.path.dirname(sys.executable)
    else:
        # Running as script
        return os.path.abspath(".")

app = Flask(__name__, 
            template_folder=get_resource_path('templates'),
            static_folder=get_resource_path('static'))

DATABASE_URL = os.environ.get('DATABASE_URL')

# Disable Flask debug output for cleaner desktop app experience
import logging
log = logging.getLogger('werkzeug')
log.setLevel(logging.ERROR)


def using_database():
    """True when DATABASE_URL is configured and psycopg is installed."""
    return bool(DATABASE_URL and psycopg)


@contextmanager
def get_db_connection():
    """Yield a PostgreSQL connection when DATABASE_URL is available."""
    if not using_database():
        raise RuntimeError('Database backend is not configured.')

    conn = psycopg.connect(DATABASE_URL)
    try:
        yield conn
    finally:
        conn.close()

# Load data
def load_data():
    data_dir = get_data_dir()
    
    with open(get_resource_path('abet_organized.json'), 'r') as f:
        abet_data = json.load(f)
    
    with open(get_resource_path('courses_organized.json'), 'r') as f:
        courses_data = json.load(f)
    
    # Mapping file is writable, so look in data directory
    mapping_path = os.path.join(data_dir, 'course_abet_mapping.json')
    
    # If mapping file doesn't exist in data dir, copy from resources
    if not os.path.exists(mapping_path):
        import shutil
        shutil.copy(get_resource_path('course_abet_mapping.json'), mapping_path)
    
    with open(mapping_path, 'r') as f:
        mapping_data = json.load(f)
    
    return abet_data, courses_data, mapping_data


def init_db_schema():
    """Create mapping table if needed and seed from JSON once."""
    if not using_database():
        return

    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS course_outcome_mappings (
                    outcome_id TEXT NOT NULL,
                    course_code TEXT NOT NULL,
                    created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    PRIMARY KEY (outcome_id, course_code)
                )
                """
            )

            # Data fix: CE-CAE.CORE.1.4 was duplicated in the source catalog.
            # Keep backward compatibility by cloning existing rows to the new
            # canonical ID CE-CAE.CORE.1.5 once (idempotent via ON CONFLICT).
            cur.execute(
                """
                INSERT INTO course_outcome_mappings (outcome_id, course_code)
                SELECT 'CE-CAE.CORE.1.5', course_code
                FROM course_outcome_mappings
                WHERE outcome_id = 'CE-CAE.CORE.1.4'
                ON CONFLICT DO NOTHING
                """
            )

            cur.execute("SELECT COUNT(*) FROM course_outcome_mappings")
            mapping_count = cur.fetchone()[0]

            if mapping_count == 0:
                _, _, mapping_data = load_data()
                rows = []
                for outcome_id, course_codes in mapping_data.get('abet_to_course', {}).items():
                    for course_code in course_codes:
                        rows.append((outcome_id, course_code))

                if rows:
                    cur.executemany(
                        """
                        INSERT INTO course_outcome_mappings (outcome_id, course_code)
                        VALUES (%s, %s)
                        ON CONFLICT DO NOTHING
                        """,
                        rows,
                    )
        conn.commit()


def build_mapping_from_rows(courses_data, rows):
    """Convert DB rows to the existing mapping JSON shape expected by the UI."""
    course_to_abet = {code: [] for code in courses_data.keys()}
    abet_to_course = {}

    for outcome_id, course_code in rows:
        abet_to_course.setdefault(outcome_id, []).append(course_code)
        course_to_abet.setdefault(course_code, []).append(outcome_id)

    for outcome_id in abet_to_course:
        abet_to_course[outcome_id] = sorted(set(abet_to_course[outcome_id]))

    for course_code in course_to_abet:
        course_to_abet[course_code] = sorted(set(course_to_abet[course_code]))

    return {
        'abet_to_course': abet_to_course,
        'course_to_abet': course_to_abet,
    }


def get_mapping_data(courses_data):
    """Return mapping data from DB when configured, else from local JSON."""
    if not using_database():
        _, _, mapping_data = load_data()
        return mapping_data

    init_db_schema()

    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT outcome_id, course_code
                FROM course_outcome_mappings
                ORDER BY outcome_id, course_code
                """
            )
            rows = cur.fetchall()
    return build_mapping_from_rows(courses_data, rows)


def save_mapping_data(mapping_data):
    """Persist mapping to DB when configured, else local JSON file."""
    if not using_database():
        data_dir = get_data_dir()
        mapping_path = os.path.join(data_dir, 'course_abet_mapping.json')
        with open(mapping_path, 'w') as f:
            json.dump(mapping_data, f, indent=2)
        return

    init_db_schema()

    rows = []
    for outcome_id, course_codes in mapping_data.get('abet_to_course', {}).items():
        for course_code in course_codes:
            rows.append((outcome_id, course_code))

    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM course_outcome_mappings")
            if rows:
                cur.executemany(
                    """
                    INSERT INTO course_outcome_mappings (outcome_id, course_code)
                    VALUES (%s, %s)
                    ON CONFLICT DO NOTHING
                    """,
                    rows,
                )
        conn.commit()


def remove_mapping_row(outcome_id, course_code):
    """Remove one mapping row from the active backend."""
    if not using_database():
        _, courses_data, mapping_data = load_data()
        if course_code in mapping_data.get('course_to_abet', {}):
            if outcome_id in mapping_data['course_to_abet'][course_code]:
                mapping_data['course_to_abet'][course_code].remove(outcome_id)
        if outcome_id in mapping_data.get('abet_to_course', {}):
            if course_code in mapping_data['abet_to_course'][outcome_id]:
                mapping_data['abet_to_course'][outcome_id].remove(course_code)
                if not mapping_data['abet_to_course'][outcome_id]:
                    del mapping_data['abet_to_course'][outcome_id]
        save_mapping_data(mapping_data)
        return build_mapping_from_rows(courses_data, [
            (oid, ccode)
            for oid, course_codes in mapping_data.get('abet_to_course', {}).items()
            for ccode in course_codes
        ])

    init_db_schema()

    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                DELETE FROM course_outcome_mappings
                WHERE outcome_id = %s AND course_code = %s
                """,
                (outcome_id, course_code),
            )
        conn.commit()


def get_export_dir():
    """Choose a user-friendly export directory."""
    downloads_dir = os.path.expanduser('~/Downloads')
    if os.path.isdir(downloads_dir):
        return downloads_dir
    return get_data_dir()


def open_file_with_default_app(file_path):
    """Open a file with the platform default application."""
    try:
        if sys.platform == 'darwin':
            subprocess.Popen(['open', file_path])
        elif os.name == 'nt':
            os.startfile(file_path)
        else:
            subprocess.Popen(['xdg-open', file_path])
    except Exception:
        pass

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/abet')
def get_abet():
    """Get all ABET learning outcomes"""
    abet_data, _, _ = load_data()
    return jsonify(abet_data)

@app.route('/api/courses')
def get_courses():
    """Get all courses"""
    _, courses_data, _ = load_data()
    return jsonify(courses_data)

@app.route('/api/mapping')
def get_mapping():
    """Get course-ABET mappings"""
    _, courses_data, _ = load_data()
    mapping_data = get_mapping_data(courses_data)
    return jsonify(mapping_data)


@app.route('/api/storage-mode')
def get_storage_mode():
    """Return current persistence backend and DB connectivity status."""
    status = {
        'backend': 'database' if using_database() else 'json',
        'database_url_present': bool(DATABASE_URL),
        'psycopg_available': bool(psycopg),
        'database_connected': False,
        'mapping_rows': None,
        'error': None,
    }

    if using_database():
        try:
            init_db_schema()
            with get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute('SELECT COUNT(*) FROM course_outcome_mappings')
                    status['mapping_rows'] = cur.fetchone()[0]
            status['database_connected'] = True
        except Exception as exc:
            status['error'] = str(exc)

    return jsonify(status)

@app.route('/api/stats')
def get_stats():
    """Get statistics about coverage"""
    abet_data, courses_data, _ = load_data()
    mapping_data = get_mapping_data(courses_data)
    
    # Count total outcomes
    total_outcomes = 0
    outcomes_by_area = {}
    
    for area_code, area_info in abet_data.items():
        count = 0
        for category, subareas in area_info['categories'].items():
            for subarea, outcomes in subareas.items():
                count += len(outcomes)
        total_outcomes += count
        outcomes_by_area[area_code] = {
            'name': area_info['name'],
            'count': count
        }
    
    # Count mapped vs unmapped
    mapped_outcomes = set(mapping_data.get('abet_to_course', {}).keys())
    
    stats = {
        'total_courses': len(courses_data),
        'total_outcomes': total_outcomes,
        'mapped_outcomes': len(mapped_outcomes),
        'unmapped_outcomes': total_outcomes - len(mapped_outcomes),
        'coverage_percentage': (len(mapped_outcomes) / total_outcomes * 100) if total_outcomes > 0 else 0,
        'outcomes_by_area': outcomes_by_area
    }
    
    return jsonify(stats)

@app.route('/api/mapping/save', methods=['POST'])
def save_mapping_route():
    """Save course-ABET mappings"""
    new_mapping = request.json
    save_mapping_data(new_mapping)
    
    return jsonify({'success': True})

@app.route('/api/remove-mapping', methods=['POST'])
def remove_mapping():
    """Remove an outcome from a course"""
    data = request.json
    outcome_id = data.get('outcome_id')
    course_code = data.get('course_code')
    
    if not outcome_id or not course_code:
        return jsonify({'error': 'Missing outcome_id or course_code'}), 400
    
    remove_mapping_row(outcome_id, course_code)
    
    return jsonify({'success': True})

@app.route('/api/gaps')
def get_gaps():
    """Find ABET outcomes not covered by any course"""
    abet_data, courses_data, _ = load_data()
    mapping_data = get_mapping_data(courses_data)
    
    # Get all mapped outcome IDs
    mapped_outcome_ids = set(mapping_data.get('abet_to_course', {}).keys())
    
    # Find unmapped outcomes
    gaps = []
    for area_code, area_info in abet_data.items():
        for category, subareas in area_info['categories'].items():
            for subarea, outcomes in subareas.items():
                for outcome in outcomes:
                    outcome_id = f"{area_code}.{category}.{subarea}.{outcome['id']}"
                    if outcome_id not in mapped_outcome_ids:
                        gaps.append({
                            'area_code': area_code,
                            'area_name': area_info['name'],
                            'category': category,
                            'subarea': subarea,
                            'id': outcome['id'],
                            'outcome': outcome['outcome'],
                            'full_id': outcome_id
                        })
    
    return jsonify(gaps)


@app.route('/api/export/coverage-matrix', methods=['POST'])
def export_coverage_matrix():
    """Export full ABET-to-course coverage matrix to an Excel workbook."""
    abet_data, courses_data, _ = load_data()
    mapping_data = get_mapping_data(courses_data)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Coverage Matrix'

    sorted_courses = sorted(courses_data.values(), key=lambda course: course['code'])
    headers = ['ABET Requirement'] + [course['code'] for course in sorted_courses]
    worksheet.append(headers)

    header_fill = PatternFill(fill_type='solid', start_color='DCE6F1', end_color='DCE6F1')
    header_font = Font(bold=True)
    covered_fill = PatternFill(fill_type='solid', start_color='E2F0D9', end_color='E2F0D9')
    section_fill = PatternFill(fill_type='solid', start_color='FFF2CC', end_color='FFF2CC')

    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    abet_to_course = mapping_data.get('abet_to_course', {})
    worksheet.sheet_properties.outlinePr.summaryBelow = False

    current_row = 2
    for area_code in sorted(abet_data.keys()):
        area_info = abet_data[area_code]
        detail_rows = []
        section_has_coverage = {course['code']: False for course in sorted_courses}

        for category, subareas in area_info['categories'].items():
            for subarea, outcomes in subareas.items():
                for outcome in outcomes:
                    outcome_id = f'{area_code}.{category}.{subarea}.{outcome["id"]}'
                    mapped_courses = set(abet_to_course.get(outcome_id, []))
                    for course in sorted_courses:
                        if course['code'] in mapped_courses:
                            section_has_coverage[course['code']] = True

                    detail_rows.append([
                        f'{outcome_id} - {outcome["outcome"]}',
                        *('x' if course['code'] in mapped_courses else '' for course in sorted_courses)
                    ])

        summary_row = [f'{area_code} - {area_info["name"]}']
        summary_row.extend('x' if section_has_coverage[course['code']] else '' for course in sorted_courses)
        worksheet.append(summary_row)

        for cell in worksheet[current_row]:
            cell.font = header_font
            cell.fill = section_fill
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if cell.column > 1 and cell.value == 'x':
                cell.fill = covered_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

        summary_row_index = current_row
        current_row += 1

        for detail_row in detail_rows:
            worksheet.append(detail_row)
            for index, cell in enumerate(worksheet[current_row], start=1):
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                if index > 1 and cell.value == 'x':
                    cell.fill = covered_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            worksheet.row_dimensions[current_row].outlineLevel = 1
            worksheet.row_dimensions[current_row].hidden = True
            current_row += 1

        if detail_rows:
            worksheet.row_dimensions[summary_row_index].collapsed = True

    worksheet.freeze_panes = 'B2'
    worksheet.column_dimensions['A'].width = 110
    for column_index in range(2, len(headers) + 1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = 14

    for row_index in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row_index].height = 42

    export_dir = get_export_dir()
    file_path = os.path.join(export_dir, 'abet-course-coverage-matrix.xlsx')
    workbook.save(file_path)
    open_file_with_default_app(file_path)

    return jsonify({'success': True, 'path': file_path})

if __name__ == '__main__':
    init_db_schema()
    port = int(os.environ.get('PORT', '5001'))
    app.run(host='0.0.0.0', port=port, debug=False, use_reloader=False)
